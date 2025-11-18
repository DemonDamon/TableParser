"""
复杂度分析器模块

实现表格复杂度智能分析功能，评估4大类7个维度：

**结构复杂度** (35%):
  - 合并单元格复杂度 (25%)
  - 表头层级复杂度 (10%)

**数据复杂度** (35%):
  - 公式/超链接 (15%)
  - 数据透视表 (10%)
  - 图表数量 (10%)

**代码复杂度** (20%):
  - VBA宏 (20%)

**规模复杂度** (10%):
  - 表格规模 (10%)

基于业界研究改进（参考：荷兰国家档案馆Spreadsheet-Complexity-Analyser、
Microsoft TableSense、Nanonets评估标准）
"""

import logging
from typing import Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .types import ComplexityScore, ComplexityLevel, OutputFormat
from .exceptions import ComplexityAnalysisError
from .utils.image_extractor import ImageExtractor

logger = logging.getLogger(__name__)


class ComplexityAnalyzer:
    """
    表格复杂度分析器
    
    根据多个维度评估表格复杂度，自动推荐最佳输出格式
    """
    
    def __init__(self):
        """初始化分析器"""
        self.image_extractor = ImageExtractor()
    
    # 动态权重配置（智能适应表格特征）
    
    # 基础权重：用于简单表格（无数据透视表/图表/VBA宏）
    WEIGHTS_BASE = {
        "merged_cells": 0.40,      # 合并单元格：40%（保持高权重）
        "header_depth": 0.30,      # 表头层级：30%（保持高权重）
        "data_structure": 0.20,    # 公式/超链接：20%
        "pivot_tables": 0.0,       # 数据透视表：0%（不计入）
        "charts": 0.0,             # 图表：0%（不计入）
        "vba_macros": 0.0,         # VBA宏：0%（不计入）
        "scale": 0.10,             # 表格规模：10%
    }
    
    # 高级权重：用于复杂表格（有数据透视表/图表/VBA宏）
    WEIGHTS_ADVANCED = {
        "merged_cells": 0.25,      # 合并单元格：25%
        "header_depth": 0.10,      # 表头层级：10%
        "data_structure": 0.15,    # 公式/超链接：15%
        "pivot_tables": 0.10,      # 数据透视表：10%
        "charts": 0.10,            # 图表：10%
        "vba_macros": 0.20,        # VBA宏：20%
        "scale": 0.10,             # 表格规模：10%
    }
    
    # 复杂度等级阈值
    THRESHOLDS = {
        "simple": 30,  # 0-30分：简单
        "medium": 60,  # 31-60分：中等
        # 61-100分：复杂
    }
    
    def analyze(self, workbook: Workbook) -> ComplexityScore:
        """
        分析表格复杂度
        
        Args:
            workbook: openpyxl Workbook对象
            
        Returns:
            ComplexityScore对象
            
        Raises:
            ComplexityAnalysisError: 分析失败时抛出
        """
        try:
            logger.info(f"开始分析表格复杂度...")
            
            # 分析所有sheet
            scores = {
                "merged_cells": 0.0,
                "header_depth": 0.0,
                "data_structure": 0.0,
                "pivot_tables": 0.0,
                "charts": 0.0,
                "vba_macros": 0.0,
                "scale": 0.0,
            }
            
            details = {
                "sheets_count": len(workbook.sheetnames),
                "total_rows": 0,
                "total_cols": 0,
                "merged_cells_count": 0,
                "has_formulas": False,
                "has_hyperlinks": False,
                "pivot_tables_count": 0,
                "charts_count": 0,
                "has_vba_macros": False,
                "images_count": 0,  # 图片数量
            }
            
            # 检测VBA宏（工作簿级别）
            vba_score = self._calculate_vba_macros_score(workbook)
            scores["vba_macros"] = vba_score
            details["has_vba_macros"] = vba_score > 0
            
            # 统计图片数量（工作簿级别）
            details["images_count"] = self.image_extractor.count_images(workbook)
            
            # 遍历所有sheet，取最大值
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # 计算各维度得分
                merged_score = self._calculate_merged_cells_score(sheet)
                header_score = self._calculate_header_depth_score(sheet)
                structure_score = self._calculate_data_structure_score(sheet)
                pivot_score = self._calculate_pivot_tables_score(sheet)
                chart_score = self._calculate_charts_score(sheet)
                scale_score = self._calculate_scale_score(sheet)
                
                # 取各维度的最大值（最复杂的sheet决定整体复杂度）
                scores["merged_cells"] = max(scores["merged_cells"], merged_score)
                scores["header_depth"] = max(scores["header_depth"], header_score)
                scores["data_structure"] = max(scores["data_structure"], structure_score)
                scores["pivot_tables"] = max(scores["pivot_tables"], pivot_score)
                scores["charts"] = max(scores["charts"], chart_score)
                scores["scale"] = max(scores["scale"], scale_score)
                
                # 累计统计信息
                details["total_rows"] += sheet.max_row
                details["total_cols"] = max(details["total_cols"], sheet.max_column)
                details["merged_cells_count"] += len(sheet.merged_cells.ranges)
                
                # 累计数据透视表和图表
                if hasattr(sheet, '_pivots'):
                    details["pivot_tables_count"] += len(sheet._pivots)
                if hasattr(sheet, '_charts'):
                    details["charts_count"] += len(sheet._charts)
            
            # 计算总分和等级
            total_score, level, recommended_format = self._calculate_total_score(scores)
            
            # 构建ComplexityScore对象
            complexity_score = ComplexityScore(
                merged_cells_score=scores["merged_cells"],
                header_depth_score=scores["header_depth"],
                data_structure_score=scores["data_structure"],
                scale_score=scores["scale"],
                total_score=total_score,
                level=level,
                recommended_format=recommended_format,
                details=details,
            )
            
            logger.info(
                f"✅ 复杂度分析完成: {level} (得分: {total_score:.1f}), "
                f"推荐格式: {recommended_format}"
            )
            
            return complexity_score
            
        except Exception as e:
            raise ComplexityAnalysisError(f"复杂度分析失败: {e}") from e
    
    def _calculate_merged_cells_score(self, sheet: Worksheet) -> float:
        """
        计算合并单元格复杂度（权重40%）
        
        评分规则：
        - 无合并单元格: 0分
        - 合并比例 < 5%: 20分
        - 合并比例 5-15%: 50分
        - 合并比例 > 15%: 80分
        - 存在跨行跨列复杂合并: +20分
        """
        merged_ranges = sheet.merged_cells.ranges
        
        if not merged_ranges:
            return 0.0
        
        total_cells = sheet.max_row * sheet.max_column
        if total_cells == 0:
            return 0.0
        
        # 计算合并单元格占比
        merged_count = sum(
            (r.max_row - r.min_row + 1) * (r.max_col - r.min_col + 1)
            for r in merged_ranges
        )
        merge_ratio = merged_count / total_cells
        
        # 基础得分
        if merge_ratio < 0.05:
            score = 20.0
        elif merge_ratio < 0.15:
            score = 50.0
        else:
            score = 80.0
        
        # 检测复杂合并（既跨行又跨列）
        has_complex_merge = any(
            (r.max_row - r.min_row > 0) and (r.max_col - r.min_col > 0)
            for r in merged_ranges
        )
        
        if has_complex_merge:
            score = min(100.0, score + 20.0)
        
        logger.debug(
            f"合并单元格得分: {score:.1f} "
            f"(合并比例: {merge_ratio*100:.1f}%, 复杂合并: {has_complex_merge})"
        )
        return score
    
    def _calculate_header_depth_score(self, sheet: Worksheet) -> float:
        """
        计算表头层级复杂度（权重30%）
        
        评分规则：
        - 单行表头: 0分
        - 2级表头: 30分
        - 3级表头: 60分
        - 4级及以上: 100分
        """
        merged_ranges = sheet.merged_cells.ranges
        
        if not merged_ranges:
            return 0.0  # 无合并，可能是单级表头
        
        # 分析前5行的合并模式（通常表头在前几行）
        header_rows = min(5, sheet.max_row)
        merged_in_header = [
            r for r in merged_ranges
            if r.min_row <= header_rows
        ]
        
        if not merged_in_header:
            return 0.0
        
        # 检测跨行合并（多级表头特征）
        max_row_span = max(
            (r.max_row - r.min_row + 1) for r in merged_in_header
        )
        
        # 根据最大跨行数判断表头层级
        if max_row_span == 1:
            score = 0.0  # 只有横向合并，单级表头
        elif max_row_span == 2:
            score = 30.0  # 2级表头
        elif max_row_span == 3:
            score = 60.0  # 3级表头
        else:
            score = 100.0  # 4级及以上
        
        logger.debug(f"表头层级得分: {score:.1f} (最大跨行: {max_row_span})")
        return score
    
    def _calculate_data_structure_score(self, sheet: Worksheet) -> float:
        """
        计算数据结构复杂度（权重20%）
        
        评分规则：
        - 纯文本数据: 0分
        - 包含公式: 30分
        - 包含超链接: 20分
        - 包含富文本/样式: 30分
        """
        score = 0.0
        
        # 采样检查（避免遍历大表）
        max_sample_rows = min(20, sheet.max_row)
        
        has_formula = False
        has_hyperlink = False
        has_rich_text = False
        
        for row in sheet.iter_rows(max_row=max_sample_rows):
            for cell in row:
                # 检测公式
                if cell.data_type == 'f':
                    has_formula = True
                
                # 检测超链接
                if cell.hyperlink:
                    has_hyperlink = True
                
                # 检测富文本（简化检测：粗体）
                if hasattr(cell, 'font') and cell.font and cell.font.bold:
                    has_rich_text = True
                
                # 提前退出（所有特征都检测到了）
                if has_formula and has_hyperlink and has_rich_text:
                    break
        
        # 累加得分
        if has_formula:
            score += 30.0
        if has_hyperlink:
            score += 20.0
        if has_rich_text:
            score += 30.0
        
        logger.debug(
            f"数据结构得分: {score:.1f} "
            f"(公式: {has_formula}, 超链接: {has_hyperlink}, 富文本: {has_rich_text})"
        )
        return min(100.0, score)
    
    def _calculate_scale_score(self, sheet: Worksheet) -> float:
        """
        计算规模复杂度（权重10%）
        
        评分规则：
        - 小表 (<100单元格): 0分
        - 中表 (100-1000): 20分
        - 大表 (1000-10000): 50分
        - 超大表 (>10000): 80分
        """
        total_cells = sheet.max_row * sheet.max_column
        
        if total_cells < 100:
            score = 0.0
        elif total_cells < 1000:
            score = 20.0
        elif total_cells < 10000:
            score = 50.0
        else:
            score = 80.0
        
        logger.debug(f"规模得分: {score:.1f} (单元格数: {total_cells})")
        return score
    
    def _calculate_pivot_tables_score(self, sheet: Worksheet) -> float:
        """
        计算数据透视表复杂度（权重10%）
        
        评分规则：
        - 无数据透视表: 0分
        - 1个数据透视表: 40分
        - 2-3个: 70分
        - 4个及以上: 100分
        """
        try:
            # 检测数据透视表
            pivot_count = 0
            if hasattr(sheet, '_pivots'):
                pivot_count = len(sheet._pivots)
            
            if pivot_count == 0:
                score = 0.0
            elif pivot_count == 1:
                score = 40.0
            elif pivot_count <= 3:
                score = 70.0
            else:
                score = 100.0
            
            logger.debug(f"数据透视表得分: {score:.1f} (数量: {pivot_count})")
            return score
        except Exception as e:
            logger.debug(f"数据透视表检测失败: {e}")
            return 0.0
    
    def _calculate_charts_score(self, sheet: Worksheet) -> float:
        """
        计算图表复杂度（权重10%）
        
        评分规则：
        - 无图表: 0分
        - 1-2个图表: 30分
        - 3-5个: 60分
        - 6个及以上: 100分
        """
        try:
            # 检测图表
            chart_count = 0
            if hasattr(sheet, '_charts'):
                chart_count = len(sheet._charts)
            
            if chart_count == 0:
                score = 0.0
            elif chart_count <= 2:
                score = 30.0
            elif chart_count <= 5:
                score = 60.0
            else:
                score = 100.0
            
            logger.debug(f"图表得分: {score:.1f} (数量: {chart_count})")
            return score
        except Exception as e:
            logger.debug(f"图表检测失败: {e}")
            return 0.0
    
    def _calculate_vba_macros_score(self, workbook: Workbook) -> float:
        """
        计算VBA宏复杂度（权重20%）
        
        评分规则：
        - 无VBA宏: 0分
        - 有VBA宏: 100分（代码复杂度极高）
        """
        try:
            # 检测VBA项目
            has_macros = False
            
            # openpyxl的workbook对象有vba_archive属性表示存在宏
            if hasattr(workbook, 'vba_archive') and workbook.vba_archive is not None:
                has_macros = True
            
            score = 100.0 if has_macros else 0.0
            
            logger.debug(f"VBA宏得分: {score:.1f} (存在: {has_macros})")
            return score
        except Exception as e:
            logger.debug(f"VBA宏检测失败: {e}")
            return 0.0
    
    def _calculate_total_score(
        self, scores: dict[str, float]
    ) -> tuple[float, ComplexityLevel, OutputFormat]:
        """
        计算综合得分并确定复杂度等级（动态权重）
        
        Args:
            scores: 各维度得分
            
        Returns:
            (总分, 复杂度等级, 推荐格式)
        """
        # 🎯 动态权重选择：检测高级特征是否存在
        has_advanced_features = (
            scores["pivot_tables"] > 0 or 
            scores["charts"] > 0 or 
            scores["vba_macros"] > 0
        )
        
        if has_advanced_features:
            # 使用高级权重（有数据透视表/图表/VBA宏）
            weights = self.WEIGHTS_ADVANCED
            weight_type = "高级权重"
        else:
            # 使用基础权重（纯结构复杂度表格）
            weights = self.WEIGHTS_BASE
            weight_type = "基础权重"
        
        # 加权求和
        total = sum(
            scores[key] * weights[key]
            for key in weights.keys()
        )
        
        logger.debug(
            f"使用{weight_type}计算总分: {total:.1f} "
            f"(数据透视表: {scores['pivot_tables']:.0f}, "
            f"图表: {scores['charts']:.0f}, "
            f"VBA宏: {scores['vba_macros']:.0f})"
        )
        
        # 确定等级和推荐格式
        if total <= self.THRESHOLDS["simple"]:
            level: ComplexityLevel = "simple"
            recommended: OutputFormat = "markdown"
        elif total <= self.THRESHOLDS["medium"]:
            level = "medium"
            recommended = "markdown"  # 可选HTML
        else:
            level = "complex"
            recommended = "html"
        
        return total, level, recommended

