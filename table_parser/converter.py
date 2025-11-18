"""
格式转换器模块

实现Workbook到Markdown和HTML的转换功能
"""

from html import escape
import logging
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .exceptions import ConversionError
from .utils.cell_utils import format_cell_value
from .utils.style_extractor import StyleExtractor
from .utils.formula_analyzer import FormulaAnalyzer
from .utils.text_formatter import TextFormatter
from .utils.rich_text_parser import RichTextParser

logger = logging.getLogger(__name__)


class FormatConverter:
    """
    格式转换器
    
    支持将Workbook转换为Markdown或HTML格式
    """
    
    def __init__(self):
        """初始化转换器"""
        self.style_extractor = StyleExtractor()
        self.formula_analyzer = FormulaAnalyzer()
        self.text_formatter = TextFormatter()
        self.rich_text_parser = RichTextParser()
        self.current_excel_path = None  # 当前处理的Excel文件路径
        self.string_index_mapping = {}  # 单元格坐标到字符串索引的映射
    
    def to_markdown(
        self,
        workbook: Workbook,
        include_empty_rows: bool = False,
        **options
    ) -> str:
        """
        转换为Markdown格式
        
        策略：
        - 每个sheet转换为独立表格
        - 使用pandas的to_markdown()
        - 合并单元格会被展开（取第一个值）
        - 添加sheet名称作为标题
        
        Args:
            workbook: openpyxl Workbook对象
            include_empty_rows: 是否包含空行
            **options: 其他选项
            
        Returns:
            Markdown格式字符串
            
        Raises:
            ConversionError: 转换失败时抛出
        """
        try:
            logger.info("开始转换为Markdown格式...")
            results = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # 提取数据
                data = []
                for row in sheet.iter_rows(values_only=True):
                    # 过滤全空行
                    if not include_empty_rows and all(v is None or str(v).strip() == "" for v in row):
                        continue
                    data.append(row)
                
                if not data:
                    logger.warning(f"Sheet '{sheet_name}' 没有数据，跳过")
                    continue
                
                # 转换为DataFrame
                if len(data) > 1:
                    df = pd.DataFrame(data[1:], columns=data[0])
                else:
                    df = pd.DataFrame(data)
                
                # 添加sheet标题（如果不是默认名称）
                if sheet_name.lower() not in ["sheet", "sheet1", "data"]:
                    results.append(f"## {sheet_name}\n")
                
                # 转换为markdown
                try:
                    md_table = df.to_markdown(index=False)
                    results.append(md_table)
                    results.append("\n")
                    logger.debug(f"Sheet '{sheet_name}' 转换完成")
                except Exception as e:
                    logger.warning(f"Sheet '{sheet_name}' 转换失败: {e}")
                    continue
            
            result = "\n".join(results)
            logger.info(f"✅ Markdown转换完成，共 {len(workbook.sheetnames)} 个sheet")
            return result
            
        except Exception as e:
            raise ConversionError(f"Markdown转换失败: {e}") from e
    
    def to_html(
        self,
        workbook: Workbook,
        chunk_rows: int = 256,
        preserve_styles: bool = False,
        include_empty_rows: bool = False,
        excel_path: Optional[str] = None,
        **options
    ) -> list[str]:
        """
        转换为HTML格式
        
        支持：
        - 合并单元格（rowspan/colspan）
        - 多级表头
        - 分块处理大表
        - 可选样式保留
        
        Args:
            workbook: openpyxl Workbook对象
            chunk_rows: 每个HTML表格的最大行数（分块处理）
            preserve_styles: 是否保留样式（暂未实现）
            include_empty_rows: 是否包含空行
            **options: 其他选项
            
        Returns:
            HTML字符串列表（每个元素是一个表格）
            
        Raises:
            ConversionError: 转换失败时抛出
        """
        try:
            logger.info("开始转换为HTML格式...")
            
            # 如果提供了Excel路径，解析sharedStrings以获取富文本格式
            if excel_path:
                self.current_excel_path = excel_path
                self.string_index_mapping = self.rich_text_parser.get_cell_string_index_mapping(excel_path)
                logger.info(f"✅ 解析了 {len(self.string_index_mapping)} 个单元格的富文本映射")
            
            html_chunks = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                try:
                    rows = list(sheet.rows)
                except Exception as e:
                    logger.warning(f"Sheet '{sheet_name}' 无法访问行数据: {e}，跳过")
                    continue
                
                if not rows:
                    logger.warning(f"Sheet '{sheet_name}' 没有数据，跳过")
                    continue
                
                # 处理表头
                header_html = self._build_header_row(rows[0])
                
                # 分块处理数据行
                for chunk_idx in range((len(rows) - 1) // chunk_rows + 1):
                    start_row = 1 + chunk_idx * chunk_rows
                    end_row = min(1 + (chunk_idx + 1) * chunk_rows, len(rows))
                    
                    html = f'<table>\n<caption>{escape(sheet_name)}</caption>\n'
                    html += '<thead>\n' + header_html + '</thead>\n'
                    html += '<tbody>\n'
                    
                    # 处理数据行
                    for row in rows[start_row:end_row]:
                        # 过滤全空行
                        if not include_empty_rows and all(
                            c.value is None or str(c.value).strip() == "" for c in row
                        ):
                            continue
                        
                        html += self._build_data_row(row, sheet, preserve_styles)
                    
                    html += '</tbody>\n</table>\n'
                    html_chunks.append(html)
                    
                    logger.debug(
                        f"Sheet '{sheet_name}' chunk {chunk_idx + 1} 完成 "
                        f"(行 {start_row}-{end_row})"
                    )
            
            logger.info(f"✅ HTML转换完成，共 {len(html_chunks)} 个表格块")
            return html_chunks
            
        except Exception as e:
            raise ConversionError(f"HTML转换失败: {e}") from e
    
    def _build_header_row(self, header_row) -> str:
        """构建HTML表头行"""
        html = "<tr>"
        for cell in header_row:
            value = format_cell_value(cell.value)
            html += f"<th>{escape(value)}</th>"
        html += "</tr>\n"
        return html
    
    def _build_data_row(self, data_row, sheet: Worksheet, preserve_styles: bool = False) -> str:
        """
        构建HTML数据行（支持合并单元格、样式、富文本）
        
        Args:
            data_row: 数据行
            sheet: 工作表对象
            preserve_styles: 是否保留样式
        """
        html = "<tr>"
        
        for cell in data_row:
            # 检查是否在合并区域中，以及是否为起始单元格
            merged_info = self._get_merge_info(cell, sheet)
            
            if merged_info == "skip":
                # 合并区域的非起始单元格，跳过
                continue
            
            # 构建单元格内容
            cell_content = self._format_cell_content(cell, output_format="html")
            
            # 构建属性字符串
            attrs = ""
            
            # 添加合并属性
            if merged_info and merged_info != "skip":
                attrs += merged_info
            
            # 添加样式属性
            if preserve_styles:
                style = self.style_extractor.get_cell_html_style(cell)
                if style:
                    attrs += f' style="{style}"'
            
            html += f"<td{attrs}>{cell_content}</td>"
        
        html += "</tr>\n"
        return html
    
    def _format_cell_content(self, cell, output_format: str = "html", show_formulas: bool = False) -> str:
        """
        格式化单元格内容（支持富文本、上下标、公式）
        
        Args:
            cell: 单元格对象
            output_format: 输出格式（html/markdown）
            show_formulas: 是否显示公式（而非计算结果）
            
        Returns:
            格式化后的内容
        """
        # 🔢 优先处理公式单元格
        if cell.data_type == 'f' and show_formulas:
            formula_str = str(cell.value) if cell.value else ""
            if not formula_str.startswith('='):
                formula_str = '=' + formula_str
            
            if output_format == "html":
                return f'<code>{escape(formula_str)}</code>'
            else:
                return f'`{formula_str}`'
        
        # 🎯 优先检查XML富文本（sharedStrings中的上下标等）
        if self.current_excel_path and cell.coordinate in self.string_index_mapping:
            string_idx = self.string_index_mapping[cell.coordinate]
            rich_text_parts = self.rich_text_parser.get_cell_rich_text(
                self.current_excel_path, 
                string_idx
            )
            
            if rich_text_parts and len(rich_text_parts) > 1:
                # 有富文本格式（上下标等）
                if output_format == "html":
                    return self.rich_text_parser.format_rich_text_to_html(rich_text_parts)
                else:
                    # Markdown简化处理
                    return ''.join(text for text, _ in rich_text_parts)
        
        # 检查openpyxl的富文本（备用）
        style_info = self.style_extractor.extract_cell_style(cell)
        
        if style_info["rich_text_parts"]:
            # 富文本格式（含上下标）
            if output_format == "html":
                return self.style_extractor.format_rich_text_to_html(style_info["rich_text_parts"])
            else:
                return self.style_extractor.format_rich_text_to_markdown(style_info["rich_text_parts"])
        else:
            # 普通文本
            value = format_cell_value(cell.value)
            
            if output_format == "html":
                # 🎯 先转换Unicode上下标字符（H₂O → H<sub>2</sub>O）
                # 注意：必须在escape之前转换，否则<sub>等标签会被转义
                formatted = self.text_formatter.convert_unicode_scripts_to_html(value)
                
                # 如果整个单元格是上标/下标格式
                if style_info.get("is_superscript"):
                    formatted = f"<sup>{formatted}</sup>"
                elif style_info.get("is_subscript"):
                    formatted = f"<sub>{formatted}</sub>"
                
                return formatted
            else:
                # Markdown
                return value
    
    def _get_merge_info(self, cell, sheet: Worksheet) -> Optional[str]:
        """
        获取单元格的合并信息
        
        Returns:
            None: 非合并单元格
            "skip": 合并区域的非起始单元格（需要跳过）
            str: rowspan/colspan属性字符串
        """
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 检查是否为起始单元格（左上角）
                if (cell.row == merged_range.min_row and 
                    cell.column == merged_range.min_col):
                    # 是起始单元格，计算rowspan和colspan
                    row_span = merged_range.max_row - merged_range.min_row + 1
                    col_span = merged_range.max_col - merged_range.min_col + 1
                    
                    attrs = ""
                    if row_span > 1:
                        attrs += f' rowspan="{row_span}"'
                    if col_span > 1:
                        attrs += f' colspan="{col_span}"'
                    
                    return attrs
                else:
                    # 不是起始单元格，需要跳过
                    return "skip"
        
        # 不在任何合并区域中
        return None
    
    def get_workbook_metadata(self, workbook: Workbook) -> dict:
        """
        获取Workbook元数据
        
        Args:
            workbook: openpyxl Workbook对象
            
        Returns:
            元数据字典
        """
        metadata = {
            "sheets": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "total_rows": 0,
            "total_cols": 0,
            "merged_cells_count": 0,
        }
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            metadata["total_rows"] += sheet.max_row
            metadata["total_cols"] = max(metadata["total_cols"], sheet.max_column)
            metadata["merged_cells_count"] += len(sheet.merged_cells.ranges)
        
        return metadata

