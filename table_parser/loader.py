"""
文件加载器模块

实现三层容错机制的文件加载功能：
1. 优先使用openpyxl（支持完整的Excel特性）
2. 失败时降级到pandas（兼容性更好）
3. pandas失败时尝试calamine引擎
4. 自动识别CSV并转换为Workbook
"""

from io import BytesIO
from pathlib import Path
from typing import Union
import logging

import pandas as pd
from openpyxl import Workbook, load_workbook

from .exceptions import FileLoadError, UnsupportedFileTypeError
from .utils.cell_utils import clean_string
from .utils.encoding_utils import detect_encoding

logger = logging.getLogger(__name__)


class FileLoader:
    """
    文件加载器
    
    采用多引擎容错机制加载Excel和CSV文件
    """
    
    # Excel文件头特征
    EXCEL_SIGNATURES = {
        b"PK\x03\x04",  # .xlsx (ZIP格式)
        b"\xd0\xcf\x11\xe0",  # .xls (OLE格式)
    }
    
    def load(self, file_like: Union[str, Path, bytes]) -> Workbook:
        """
        加载文件并返回openpyxl Workbook对象
        
        Args:
            file_like: 文件路径、Path对象或二进制内容
            
        Returns:
            openpyxl Workbook对象
            
        Raises:
            FileLoadError: 加载失败时抛出
            UnsupportedFileTypeError: 不支持的文件类型
        """
        # 统一转换为BytesIO
        file_obj = self._to_bytes_io(file_like)
        
        # 判断文件类型
        if self._is_excel_file(file_obj):
            return self._load_excel(file_obj)
        else:
            return self._load_csv(file_obj)
    
    def _to_bytes_io(self, file_like: Union[str, Path, bytes]) -> BytesIO:
        """将输入转换为BytesIO对象"""
        if isinstance(file_like, (str, Path)):
            file_path = Path(file_like)
            try:
                with open(file_path, "rb") as f:
                    content = f.read()
                logger.info(f"成功读取文件: {file_path} ({len(content)} bytes)")
                return BytesIO(content)
            except Exception as e:
                raise FileLoadError(f"读取文件失败: {file_path}") from e
        
        elif isinstance(file_like, bytes):
            logger.info(f"接收到二进制内容 ({len(file_like)} bytes)")
            return BytesIO(file_like)
        
        else:
            raise UnsupportedFileTypeError(f"不支持的输入类型: {type(file_like)}")
    
    def _is_excel_file(self, file_obj: BytesIO) -> bool:
        """判断是否为Excel文件"""
        file_obj.seek(0)
        file_head = file_obj.read(4)
        file_obj.seek(0)
        
        is_excel = any(file_head.startswith(sig) for sig in self.EXCEL_SIGNATURES)
        logger.debug(f"文件类型判断: {'Excel' if is_excel else 'CSV'}")
        return is_excel
    
    def _load_excel(self, file_obj: BytesIO) -> Workbook:
        """
        加载Excel文件（三层容错）
        
        1. 尝试openpyxl
        2. 失败则尝试pandas默认引擎
        3. 再失败则尝试pandas + calamine引擎
        """
        # 第一层：openpyxl
        try:
            logger.debug("尝试使用openpyxl加载...")
            file_obj.seek(0)
            wb = load_workbook(file_obj, data_only=True)
            logger.info(f"✅ openpyxl加载成功，包含 {len(wb.sheetnames)} 个sheet")
            return wb
        except Exception as e:
            logger.warning(f"openpyxl加载失败: {e}")
        
        # 第二层：pandas默认引擎
        try:
            logger.debug("尝试使用pandas默认引擎加载...")
            file_obj.seek(0)
            dfs = pd.read_excel(file_obj, sheet_name=None)
            wb = self._dataframes_to_workbook(dfs)
            logger.info(f"✅ pandas默认引擎加载成功，包含 {len(wb.sheetnames)} 个sheet")
            return wb
        except Exception as e:
            logger.warning(f"pandas默认引擎加载失败: {e}")
        
        # 第三层：pandas + calamine引擎
        try:
            logger.debug("尝试使用pandas + calamine引擎加载...")
            file_obj.seek(0)
            df = pd.read_excel(file_obj, engine="calamine")
            wb = self._dataframe_to_workbook(df)
            logger.info(f"✅ pandas + calamine引擎加载成功")
            return wb
        except Exception as e_calamine:
            logger.error(f"pandas + calamine引擎加载失败: {e_calamine}")
            raise FileLoadError(
                f"所有Excel加载引擎都失败了。请检查文件是否损坏。"
            ) from e_calamine
    
    def _load_csv(self, file_obj: BytesIO) -> Workbook:
        """加载CSV文件"""
        try:
            logger.debug("尝试加载CSV文件...")
            file_obj.seek(0)
            content = file_obj.read()
            
            # 检测编码
            encoding = detect_encoding(content)
            logger.debug(f"使用编码: {encoding}")
            
            # 使用pandas读取CSV
            file_obj.seek(0)
            df = pd.read_csv(file_obj, encoding=encoding)
            
            # 转换为Workbook
            wb = self._dataframe_to_workbook(df)
            logger.info(f"✅ CSV文件加载成功，{len(df)} 行 × {len(df.columns)} 列")
            return wb
            
        except Exception as e:
            raise FileLoadError(f"CSV文件加载失败: {e}") from e
    
    def _dataframe_to_workbook(self, df: pd.DataFrame, sheet_name: str = "Data") -> Workbook:
        """将单个DataFrame转换为Workbook"""
        # 清理数据
        df = df.apply(lambda col: col.map(clean_string))
        
        # 创建Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 写入表头
        for col_num, column_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)
        
        # 写入数据
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        logger.debug(f"DataFrame转Workbook完成: {len(df)} 行 × {len(df.columns)} 列")
        return wb
    
    def _dataframes_to_workbook(self, dfs: dict[str, pd.DataFrame]) -> Workbook:
        """将多个DataFrame（多sheet）转换为Workbook"""
        if len(dfs) == 1:
            # 只有一个sheet，使用简化方法
            sheet_name, df = next(iter(dfs.items()))
            return self._dataframe_to_workbook(df, sheet_name)
        
        # 多个sheet
        wb = Workbook()
        # 删除默认sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # 创建每个sheet
        for sheet_name, df in dfs.items():
            # 清理数据
            df = df.apply(lambda col: col.map(clean_string))
            
            ws = wb.create_sheet(title=sheet_name)
            
            # 写入表头
            for col_num, column_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_num, value=column_name)
            
            # 写入数据
            for row_num, row_data in enumerate(df.values, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
        
        logger.debug(f"多sheet DataFrame转Workbook完成: {len(dfs)} 个sheet")
        return wb

